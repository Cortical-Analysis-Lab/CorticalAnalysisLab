"""Shared helpers for the summer-research catalog tooling."""

from __future__ import annotations

import csv
import hashlib
import json
import re
import sqlite3
import unicodedata
from pathlib import Path
from urllib.parse import urlparse

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_DB = ROOT / "database" / "research_opportunities.sqlite"
SCHEMA = ROOT / "schema" / "schema.sql"
IMPORTER_VERSION = "1.2.1"
NA_VALUES = {"n/a", "na", "not applicable"}

NON_AUTHORITATIVE_EVIDENCE_HOSTS = {
    "facebook.com", "instagram.com", "linkedin.com", "reddit.com", "tiktok.com",
    "twitter.com", "x.com", "youtube.com", "pathwaystoscience.org",
}

CATEGORIES = [
    ("biomedical-health", "Biomedical & Health", "Human health, medicine, public health, and biomedical research"),
    ("life-sciences", "Life Sciences", "Biology, molecular and cellular sciences, ecology, and related life sciences"),
    ("neuroscience-cognitive", "Neuroscience & Cognitive Science", "Neural systems, behavior, cognition, and brain health"),
    ("computer-data-ai", "Computer Science / Data / AI", "Computing, data science, machine learning, and artificial intelligence"),
    ("mathematics-statistics", "Mathematics & Statistics", "Pure and applied mathematics, statistics, and quantitative modeling"),
    ("physics-astronomy", "Physics & Astronomy", "Physics, astronomy, and space sciences"),
    ("chemistry-materials", "Chemistry & Materials", "Chemistry, chemical science, and materials research"),
    ("engineering", "Engineering", "Engineering research and design across disciplines"),
    ("earth-environment-ocean", "Earth / Environment / Ocean", "Earth systems, environmental science, climate, and ocean research"),
    ("social-behavioral", "Social / Behavioral Sciences", "Psychology, sociology, education, economics, and related social sciences"),
    ("humanities-arts", "Humanities & Arts", "Humanities, arts, languages, and humanistic scholarship"),
    ("interdisciplinary-stem", "Interdisciplinary STEM", "Programs intentionally spanning multiple STEM disciplines"),
    ("multidisciplinary", "Multidisciplinary", "Programs spanning STEM, social sciences, humanities, or other broad areas"),
]

TAG_ALIASES = {
    "stem": ("discipline", "STEM"),
    "social sciences": ("discipline", "Social sciences"),
    "humanities": ("discipline", "Humanities"),
    "biomedical": ("discipline", "Biomedical research"),
    "neuroscience": ("discipline", "Neuroscience"),
    "computational": ("research_mode", "Computational research"),
    "data science": ("discipline", "Data science"),
    "machine learning": ("discipline", "Machine learning"),
    "field research": ("research_mode", "Field research"),
    "clinical": ("research_mode", "Clinical research"),
    "engineering": ("discipline", "Engineering"),
    "phd-oriented research": ("program_characteristic", "PhD-oriented research"),
    "international": ("program_characteristic", "International"),
    "varies by host": ("scope", "Varies by host"),
    "varies by project": ("scope", "Varies by project"),
}

RESEARCH_MODES = [
    ("wet_lab", "Wet lab", "Laboratory research involving biological, chemical, or physical materials"),
    ("computational", "Computational", "Research centered on computation, software, simulation, or data"),
    ("field", "Field research", "Research conducted primarily in natural or community settings"),
    ("clinical", "Clinical", "Research involving clinical settings, populations, or health data"),
    ("translational", "Translational", "Research connecting foundational findings to practical or clinical use"),
    ("engineering_design", "Engineering / design", "Engineering, prototyping, systems, or design work"),
    ("theoretical", "Theoretical", "Theory-focused or mathematical research"),
    ("archival", "Archival", "Research centered on archives, collections, or historical records"),
    ("qualitative", "Qualitative", "Interview, observational, textual, or other qualitative methods"),
    ("quantitative", "Quantitative", "Statistical, mathematical, or other quantitative methods"),
    ("mixed", "Mixed methods", "Programs explicitly combining multiple research modes"),
]

MODE_ALIASES = {
    "wet lab": "wet_lab",
    "wet_lab": "wet_lab",
    "computational": "computational",
    "computational research": "computational",
    "field": "field",
    "field research": "field",
    "clinical": "clinical",
    "clinical research": "clinical",
    "translational": "translational",
    "engineering/design": "engineering_design",
    "engineering design": "engineering_design",
    "theoretical": "theoretical",
    "archival": "archival",
    "qualitative": "qualitative",
    "quantitative": "quantitative",
    "mixed": "mixed",
    "mixed methods": "mixed",
}


def research_modes_from_tag(value):
    """Return modes stated explicitly in a reviewed tag; do not infer methods."""
    normalized = (text_or_none(value) or "").lower().replace("_", " ")
    modes = set()
    if "computational" in normalized or normalized in {"computation", "computing"}:
        modes.add("computational")
    if "field research" in normalized:
        modes.add("field")
    if "clinical research" in normalized:
        modes.add("clinical")
    if "translational" in normalized:
        modes.add("translational")
    if "theoretical" in normalized:
        modes.add("theoretical")
    if "archival" in normalized:
        modes.add("archival")
    if "qualitative" in normalized:
        modes.add("qualitative")
    if "quantitative" in normalized:
        modes.add("quantitative")
    if "mixed method" in normalized:
        modes.add("mixed")
    return sorted(modes)


def text_or_none(value):
    if value is None:
        return None
    value = str(value).strip()
    return value or None


def url_host(value):
    """Return a normalized URL hostname without treating it as proof of authority."""
    return (urlparse(text_or_none(value) or "").hostname or "").lower().rstrip(".")


def host_is_or_belongs_to(host, parent):
    """Match an exact host or a real subdomain boundary, never a string suffix alone."""
    return bool(host and parent and (host == parent or host.endswith("." + parent)))


def disallowed_verification_source(value):
    """Reject social and discovery-only hosts as canonical field evidence."""
    host = url_host(value)
    return any(host_is_or_belongs_to(host, blocked) for blocked in NON_AUTHORITATIVE_EVIDENCE_HOSTS)


def eligibility_source_matches_official_program(source_url, program_url):
    """Require eligibility evidence to share the reviewed program's domain family.

    Cross-domain government/network evidence must be modeled as an explicitly reviewed
    additional source rather than silently accepted by the flat seed importer.
    """
    source_host = url_host(source_url)
    program_host = url_host(program_url)
    if disallowed_verification_source(source_url):
        return False
    return host_is_or_belongs_to(source_host, program_host) or host_is_or_belongs_to(program_host, source_host)


def number_or_none(value):
    value = text_or_none(value)
    if value is None:
        return None
    if value.lower() in NA_VALUES:
        return None
    return float(value.replace(",", "").replace("$", ""))


def int_or_none(value):
    number = number_or_none(value)
    return int(number) if number is not None else None


def iso_date(value):
    value = text_or_none(value)
    if value is None:
        return None
    if value.lower() in NA_VALUES:
        return None
    return value[:10]


def slugify(value):
    value = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode()
    value = re.sub(r"[^a-zA-Z0-9]+", "-", value.lower()).strip("-")
    return value or "unknown"


def normalize_country(value):
    value = text_or_none(value)
    return {"USA": "US", "United States": "US", "Germany": "DE"}.get(value, value)


def normalize_status(value):
    value = (text_or_none(value) or "").lower()
    if "closed" in value or "filled" in value:
        return "closed"
    if "upcoming" in value or "expected" in value or "monitor" in value:
        return "upcoming"
    if "open" in value:
        return "open"
    if "active" in value:
        return "active"
    return "unknown"


def normalize_choice(value):
    value = (text_or_none(value) or "unknown").lower().replace("not specified", "unknown")
    return re.sub(r"[^a-z0-9]+", "_", value).strip("_") or "unknown"


def normalize_external(value):
    value = normalize_choice(value)
    if value in {"yes", "no", "limited"}:
        return value
    return "unknown"


def safe_bool_from_explicit(text, positive_patterns, negative_patterns=()):
    """Return only explicit True/False; ambiguity remains None."""
    lower = (text_or_none(text) or "").lower()
    if any(pattern in lower for pattern in negative_patterns):
        return 0
    if any(pattern in lower for pattern in positive_patterns):
        return 1
    return None


def load_rows(path: Path):
    if path.suffix.lower() == ".csv":
        with path.open(encoding="utf-8-sig", newline="") as handle:
            return list(csv.DictReader(handle))
    if path.suffix.lower() == ".xlsx":
        try:
            import openpyxl
        except ImportError as exc:
            raise SystemExit("XLSX import requires: pip install -r scripts/requirements.txt") from exc
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet = workbook["Programs"] if "Programs" in workbook.sheetnames else workbook.active
        values = sheet.iter_rows(values_only=True)
        headers = [str(value) if value is not None else "" for value in next(values)]
        return [dict(zip(headers, row)) for row in values if any(value is not None for value in row)]
    raise ValueError(f"Unsupported import format: {path.suffix}")


def sha256(path: Path):
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def valid_url(value):
    if value is None:
        return True
    parsed = urlparse(value)
    return parsed.scheme in {"http", "https"} and bool(parsed.netloc)


def connect(db_path=DEFAULT_DB):
    connection = sqlite3.connect(db_path)
    connection.row_factory = sqlite3.Row
    connection.execute("PRAGMA foreign_keys = ON")
    return connection


def dump_json(path, data):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
