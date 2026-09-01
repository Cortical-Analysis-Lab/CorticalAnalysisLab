#!/usr/bin/env python3
"""Package generated review CSVs into a filterable XLSX workbook."""

from __future__ import annotations

import csv
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError as exc:
    raise SystemExit("XLSX export requires: pip install -r scripts/requirements.txt") from exc

from catalog_common import ROOT

REVIEW = ROOT / "data" / "summer-research" / "review"
FILES = [
    ("Opportunities", "opportunities_review.csv"),
    ("Institutions", "institutions.csv"),
    ("Cycles", "program_cycles.csv"),
    ("Eligibility", "eligibility_rules.csv"),
    ("Sources", "sources.csv"),
    ("Verifications", "source_verifications.csv"),
]


def main():
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name, filename in FILES:
        sheet = workbook.create_sheet(sheet_name)
        with (REVIEW / filename).open(encoding="utf-8", newline="") as handle:
            for row in csv.reader(handle):
                sheet.append(row)
        sheet.freeze_panes = "D2"
        sheet.sheet_view.showGridLines = False
        for cell in sheet[1]:
            cell.fill = PatternFill("solid", fgColor="173F5F")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        sheet.row_dimensions[1].height = 34
        for column in sheet.columns:
            width = min(34, max(11, max(len(str(cell.value or "")) for cell in list(column)[:20]) + 2))
            sheet.column_dimensions[column[0].column_letter].width = width
            for cell in column:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        if sheet.max_row > 1 and sheet.max_column > 0:
            ref = f"A1:{sheet.cell(sheet.max_row, sheet.max_column).coordinate}"
            table = Table(displayName=f"{sheet_name}Table", ref=ref)
            table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False, showLastColumn=False)
            sheet.add_table(table)
    destination = REVIEW / "summer_research_catalog_review.xlsx"
    workbook.save(destination)
    print(destination)


if __name__ == "__main__":
    main()
